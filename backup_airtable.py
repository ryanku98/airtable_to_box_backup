import requests # .get()
from openpyxl import Workbook, load_workbook
from sys import exit
from datetime import datetime
import os # remove, rename, path.exists
from json import loads
from boxsdk import JWTAuth, Client
from glob import glob

last_upload_name = "last_uploaded.xlsx"
airtable_url_path = "config/airtable_url.txt"
airtable_config_path = "config/airtable_config.txt"
box_config_path = "config/box_config.json"
box_folder_id_path = "config/box_folder_id.txt"
box_last_folder_id_path = "config/box_last_folder_id.txt"
client = None
uploaded = False # flag to indicate if new version has been uploaded

def backup():
    # Configure AirTable authentication
    at_url = open(airtable_url_path, "r")
    at_api_key = open(airtable_config_path, "r")

    # Retrieve AirTable data
    res = requests.get(at_url.read(), headers={"Authorization": "Bearer " + at_api_key.read()})
    at_url.close()
    at_api_key.close()
    if res.status_code != 200:
        print("ERROR: could not connect to AirTable - status code", res.status_code, "\nExiting...")
        exit()

    # Create openpyxl workbook
    wb = Workbook()
    ws = wb.active

    # Write data into corresponding columns
    excel_header_dictionary = {}
    record_row = 2
    num_headers = 1
    for record in res.json()['records']:
        for k, v in record['fields'].items():
            if k not in excel_header_dictionary:
                # Add headers to dictionary and write to excel file if didn't previously exist
                excel_header_dictionary.update({k : num_headers})
                ws.cell(row=1, column=num_headers, value=k)
                num_headers += 1
            # Write values under appropriate column
            if isinstance(v, list): # if v is a list
                ws.cell(row=record_row, column=excel_header_dictionary[k], value=", ".join(v))
            else:
                ws.cell(row=record_row, column=excel_header_dictionary[k], value=v)
        record_row += 1

    # Create name of backup file with timestamp and save
    time_stamp = datetime.now().strftime("%Y-%m-%d--%H-%M-%S")
    file_name_template = "AirTable_Backup_"
    file_name = file_name_template + time_stamp + ".xlsx"
    wb.save(filename=file_name)

    # Login to Box
    # Import JWT auth object with config file
    sdk = JWTAuth.from_settings_file(box_config_path)
    try:
        global client
        client = Client(sdk) # authenticated client
        # If error occurs here, backup file persists and will be uploaded during next successful run
    except:
        return

    # Upload any previous backups if they failed
    backups = glob(file_name_template + "*")    # Retrieves non-uploaded backups in order as list
    backups.append(file_name)                   # Add current backup to list
    download_last()
    for backup_file in backups:
        if os.path.exists(backup_file):
            if not same_as_last(backup_file, rows=record_row, cols=num_headers):
                upload(client=client, file_name=backup_file)
            else: # Delete without upload if the same
                os.remove(backup_file)
                print("EVENT: File", backup_file, "not uploaded, identical to last")

    if uploaded: # only reupload last if it has been changed
        upload_last()

# download last_uploaded.xlsx from special folder in Box - snapshot of most recently backed-up version
def download_last():
    remove_last_local() # remove local last_uploaded.xlsx if it exists

    f_id = open(box_last_folder_id_path, "r")
    folder_id = f_id.read()
    f_id.close()

    # get file id
    items_iter = client.folder(folder_id).get_items()
    try:
        item = items_iter.next()
        file_id = item.id

        # download using file id
        box_file = client.file(file_id=file_id).get()
        output_file = open(box_file.name, "wb")
        box_file.download_to(output_file)
        print("EVENT: File", box_file.name, "downloaded from Box with ID:", file_id)
    except:
        # no items
        print("EVENT:", last_upload_name, "not found")

def upload(client, file_name):
    global uploaded
    uploaded = True
    # Get folder ID
    f_id = open(box_folder_id_path, "r")
    folder_id = f_id.read()
    f_id.close()

    # Upload backup to backup folder on Box
    try:
        box_file = client.folder(folder_id).upload(file_name)
        print("EVENT: File", file_name, "uploaded with ID:", box_file.id)
        # Set uploaded file as last_uploaded
        remove_last_local()
        os.rename(file_name, last_upload_name)
    except:
        print("EVENT: Error uploading file", file_name, "- possible name collision. Deleting...")
        os.remove(file_name)

def upload_last():
    f_id = open(box_last_folder_id_path, "r")
    folder_id = f_id.read()
    f_id.close()

    # empty last_uploaded folder to maintain only 1 object when backup() ends
    items_iter = client.folder(folder_id=folder_id).get_items()
    try:
        for item in items_iter:
            client.file(file_id=item.id).delete()
    except:
        # folder already empty
        pass

    box_file = client.folder(folder_id).upload(last_upload_name)
    print("EVENT: File", last_upload_name, "uploaded with ID:", box_file.id)
    remove_last_local()

def same_as_last(doc, rows, cols):
    if not os.path.exists(last_upload_name):
        return False
    wb1 = load_workbook(filename=last_upload_name)
    wb2 = load_workbook(filename=doc)
    ws1 = wb1.active
    ws2 = wb2.active
    for row in range(1, rows+1):
        for col in range(1, cols+1):
            if ws1.cell(row=row, column=col).value != ws2.cell(row=row, column=col).value:
                return False # Return false if found mismatched cell
    return True

def remove_last_local():
    if os.path.exists(last_upload_name):
        os.remove(last_upload_name)

if __name__ == "__main__":
    backup()
